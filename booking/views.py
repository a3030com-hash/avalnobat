
import datetime
import jdatetime
import random
import time
import logging
from django.conf import settings
from django.db import transaction, OperationalError
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from .models import DoctorProfile, DoctorAvailability, Appointment, TimeSlotException, Review
from .forms import DoctorAvailabilityForm, AppointmentBookingForm, ReviewForm
from django.urls import reverse
from django.db.models import Q
import requests
from django.db.models import Q, Avg
from django.http import HttpResponse
import pytz
import openpyxl
from .decorators import doctor_required, secretary_required


def _get_doctor_profile(user):
    """
    Helper function to get the doctor profile for a doctor or secretary.
    """
    if user.user_type == 'DOCTOR':
        return user.doctor_profile
    elif user.user_type == 'SECRETARY':
        return user.doctor
    return None

def doctor_list(request):
    """
    نمایش لیست تمام پزشکان با قابلیت جستجو.
    """
    queryset = DoctorProfile.objects.select_related('user', 'specialty').prefetch_related('availabilities').all()
    query = request.GET.get('q')

    if query:
        # Split the search query into individual, non-empty words
        search_terms = [term for term in query.split() if term]
        for term in search_terms:
            # For each term, filter the queryset cumulatively
            queryset = queryset.filter(
                Q(user__first_name__icontains=term) |
                Q(user__last_name__icontains=term) |
                Q(specialty__name__icontains=term) |
                Q(address__icontains=term)
            )

    context = {
        'doctors': queryset,
        'page_title':'لیست پزشکان'
    }
    return render(request, 'booking/doctor_list.html', context)

def doctor_detail(request, pk):
    """
    نمایش جزئیات یک پزشک خاص و تقویم نوبت‌دهی او بر اساس تاریخ شمسی.
    """
    doctor = get_object_or_404(DoctorProfile.objects.select_related('user', 'specialty'), pk=pk)
    availabilities = doctor.availabilities.filter(is_active=True)
    reviews = Review.objects.filter(appointment__doctor=doctor)
    average_rating = reviews.aggregate(Avg('rating'))['rating__avg']

    # محاسبه تقویم برای 45 روز آینده
    today = datetime.date.today()
    booking_days = doctor.booking_days

    j_to_model_weekday_map = {
        0: 5, 1: 6, 2: 0, 3: 1, 4: 2, 5: 3, 6: 4
    }

    jalali_day_names = ["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنجشنبه", "جمعه"]

    available_days = []
    for i in range(booking_days):
        current_gregorian_date = today + datetime.timedelta(days=i)
        current_jalali_date = jdatetime.date.fromgregorian(date=current_gregorian_date)
        model_weekday = j_to_model_weekday_map[current_jalali_date.weekday()]

        daily_availabilities = availabilities.filter(day_of_week=model_weekday)

        if daily_availabilities.exists():
            total_capacity = sum(da.visit_count for da in daily_availabilities)
            if total_capacity > 0:
                booked_count = Appointment.objects.filter(
                    doctor=doctor,
                    appointment_datetime__date=current_gregorian_date,
                    status__in=[1, 2, 4]
                ).count()

                if booked_count < total_capacity:
                    day_info = {
                        'date': current_gregorian_date,
                        'jalali_day_name': jalali_day_names[current_jalali_date.weekday()]
                    }
                    available_days.append(day_info)

    context = {
        'doctor': doctor,
        'available_days': available_days,
        'average_rating': average_rating,
        'page_title': f'پروفایل دکتر {doctor.user.get_full_name()}'
    }
    return render(request, 'booking/doctor_detail.html', context)

@login_required
def doctor_dashboard(request):
    """
    داشبورد پزشک برای مدیریت زمان‌بندی کاری.
    """
    if request.user.user_type == 'DOCTOR':
        doctor_profile = request.user.doctor_profile
    elif request.user.user_type == 'SECRETARY':
        return redirect('booking:daily_patients')
    else:
        return redirect('booking:doctor_list')

    if request.method == 'POST':
        form = DoctorAvailabilityForm(request.POST)
        if form.is_valid():
            availability = form.save(commit=False)
            availability.doctor = doctor_profile
            availability.save()
            return redirect('booking:doctor_dashboard')
    else:
        form = DoctorAvailabilityForm()

    availabilities = DoctorAvailability.objects.filter(doctor=doctor_profile).order_by('day_of_week', 'start_time')

    context = {
        'form': form,
        'availabilities': availabilities,
        'page_title': 'داشبورد مدیریت زمان‌بندی'
    }
    return render(request, 'booking/doctor_dashboard.html', context)

@login_required
@doctor_required
def edit_availability(request, pk):
    availability = get_object_or_404(DoctorAvailability, pk=pk, doctor=request.user.doctor_profile)
    if request.method == 'POST':
        form = DoctorAvailabilityForm(request.POST, instance=availability)
        if form.is_valid():
            form.save()
            return redirect('booking:doctor_oard')
    else:
        form = DoctorAvailabilityForm(instance=availability)

    context = {
        'form': form,
        'page_title': 'ویرایش برنامه کاری'
    }
    return render(request, 'booking/edit_availability.html', context)

@login_required
def delete_availability(request, pk):
    availability = get_object_or_404(DoctorAvailability, pk=pk, doctor=request.user.doctor_profile)
    if request.method == 'POST':
        availability.delete()
        return redirect('booking:doctor_dashboard')

    context = {
        'availability': availability,
        'page_title': 'تایید حذف برنامه کاری'
    }
    return render(request, 'booking/delete_availability.html', context)

@login_required
def toggle_availability(request, pk):
    availability = get_object_or_404(DoctorAvailability, pk=pk, doctor=request.user.doctor_profile)
    availability.is_active = not availability.is_active
    availability.save()
    return redirect('booking:doctor_dashboard')

def book_appointment(request, pk, date):
    """
    نمایش تمام ساعات (خالی و پر) و پردازش رزرو نوبت.
    """
    doctor = get_object_or_404(DoctorProfile, pk=pk)
    try:
        jalali_date = jdatetime.datetime.strptime(date, '%Y-%m-%d').date()
        target_date = jalali_date.togregorian()
    except ValueError:
        return redirect('booking:doctor_detail', pk=doctor.pk)

    day_of_week = target_date.weekday()
    availabilities = DoctorAvailability.objects.filter(doctor=doctor, day_of_week=day_of_week, is_active=True)

    if not availabilities.exists():
        return redirect('booking:doctor_detail', pk=doctor.pk)

    all_slots = []
    booked_datetimes = list(Appointment.objects.filter(
        doctor=doctor,
        appointment_datetime__date=target_date,
        status__in=[1, 2, 4]
    ).values_list('appointment_datetime', flat=True))

    canceled_slots = list(TimeSlotException.objects.filter(
        doctor=doctor,
        datetime_slot__date=target_date
    ).values_list('datetime_slot', flat=True))

    for avail in availabilities:
        duration = (datetime.datetime.combine(target_date, avail.end_time) - datetime.datetime.combine(target_date, avail.start_time))
        interval = duration / avail.visit_count if avail.visit_count > 1 else duration
        current_time_naive = datetime.datetime.combine(target_date, avail.start_time)
        for i in range(avail.visit_count):
            current_time_aware = timezone.make_aware(current_time_naive)
            if current_time_aware in booked_datetimes:
                status = 'booked'
            elif current_time_aware in canceled_slots:
                status = 'canceled'
            else:
                status = 'available'
            all_slots.append({'time': current_time_aware, 'status': status})
            current_time_naive += interval

    if request.method == 'POST':
        form = AppointmentBookingForm(request.POST)
        selected_slot_str = request.POST.get('selected_slot')

        if form.is_valid() and selected_slot_str:
            appointment_datetime = datetime.datetime.fromisoformat(selected_slot_str)

            try:
                with transaction.atomic():
                    if Appointment.objects.filter(doctor=doctor, appointment_datetime=appointment_datetime, status__in=[1, 2, 4]).exists():
                        raise ValueError("این نوبت لحظاتی پیش رزرو شد.")

                    appointment = form.save(commit=False)
                    appointment.doctor = doctor
                    appointment.appointment_datetime = appointment_datetime
                    appointment.status = 4
                    appointment.save()

                    # --- OTP & SMS Sending Logic ---
                    try:
                        AMOOT_SMS_API_TOKEN=settings.AMOOT_SMS_API_TOKEN
                        AMOOT_SMS_API_URL=settings.AMOOT_SMS_API_URL
                        otp_code = str(random.randint(100000, 999999))
                        request.session['otp_code'] = otp_code
                        request.session['pending_appointment_id'] = appointment.id
                        payload = {
                               'token': AMOOT_SMS_API_TOKEN,
                               'Mobile': appointment.patient_phone,
                               'PatternCodeID':4018,
                               'PatternValues': otp_code,
                            }
                        response=requests.post(AMOOT_SMS_API_URL,data=payload)
                    except requests.exceptions.RequestException as e:
                        print("خطا در ارسال پیامک:", e)
                    return redirect('booking:verify_appointment')

            except ValueError as e:
                error_message = str(e)
                context = {
                    'doctor': doctor, 'date': target_date, 'all_slots': all_slots,
                    'form': form, 'page_title': 'خطا در رزرو نوبت', 'error': error_message
                }
                return render(request, 'booking/book_appointment.html', context)
            except OperationalError as e:
                error_message = "سیستم در حال حاضر قادر به پردازش رزرو شما نیست. لطفاً بعداً دوباره تلاش کنید."
                if "attempt to write a readonly database" in str(e):
                    error_message = "مشکل فنی: پایگاه داده در حالت فقط-خواندنی قرار دارد و امکان ثبت نوبت جدید وجود ندارد. لطفاً به پشتیبانی اطلاع دهید."

                context = {
                    'doctor': doctor, 'date': target_date, 'all_slots': all_slots,
                    'form': form, 'page_title': 'خطای سیستمی', 'error': error_message
                }
                return render(request, 'booking/book_appointment.html', context)
    else:
        form = AppointmentBookingForm()

    context = {
        'doctor': doctor, 'date': target_date, 'all_slots': all_slots,
        'form': form
    }
    return render(request, 'booking/book_appointment.html', context)

from django.contrib.auth import get_user_model
User = get_user_model()

def verify_appointment(request):
    """
تایید شماره همراه با کد یک‌بار مصرف  session.
    """
    pending_appointment_id = request.session.get('pending_appointment_id')
    if not pending_appointment_id:
        return redirect('booking:doctor_list') # اگر نوبتی در حال رزرو نباشد

    appointment = get_object_or_404(Appointment, pk=pending_appointment_id)

    if request.method == 'POST':
        otp_from_user = request.POST.get('otp')
        otp_from_session = request.session.get('otp_code')

        if otp_from_user == otp_from_session:
            # Find or create a patient user with the phone number
            patient_user, created = User.objects.get_or_create(
                username=appointment.patient_phone,
                defaults={
                    'first_name': appointment.patient_name,
                    'user_type': 'PATIENT'
                }
            )
            # Assign the patient to the appointment
            appointment.patient = patient_user
            appointment.save()

            # Store the phone number for the payment page to pick up
            request.session['verified_patient_phone'] = appointment.patient_phone

            return redirect('booking:payment_page')
        else:
            # کد اشتباه است
            return render(request, 'booking/verify_appointment.html', {'error': 'کد وارد شده صحیح نمی‌باشد.'})

    return render(request, 'booking/verify_appointment.html', {'page_title': 'تأیید نوبت'})

MELLAT_BANK_ERRORS = {
    '0': 'تراکنش با موفقیت انجام شد',
    '11': 'شماره کارت نامعتبر است',
    '12': 'موجودی کافی نیست',
    '13': 'رمز نادرست است',
    '14': 'تعداد دفعات وارد کردن رمز بیش از حد مجاز است',
    '15': 'کارت نامعتبر است',
    '16': 'دفعات برداشت وجه بیش از حد مجاز است',
    '17': 'کاربر از انجام تراکنش منصرف شده است',
    '18': 'تاریخ انقضای کارت گذشته است',
    '19': 'مبلغ برداشت وجه بیش از حد مجاز است',
    '21': 'پذیرنده نامعتبر است',
    '23': 'خطای امنیتی رخ داده است',
    '24': 'اطلاعات کاربری پذیرنده نامعتبر است',
    '25': 'مبلغ نامعتبر است',
    '31': 'پاسخ نامعتبر است',
    '32': 'فرمت اطلاعات وارد شده صحیح نمی باشد',
    '33': 'حساب نامعتبر است',
    '34': 'خطای سیستمی',
    '35': 'تاریخ نامعتبر است',
    '41': 'شماره درخواست تکراری است',
    '42': 'تراکنش Sale یافت نشد',
    '43': 'قبلا درخواست Verify داده شده است',
    '44': 'درخواست Verfiy یافت نشد',
    '45': 'تراکنش Settle (تسویه) شده است',
    '46': 'تراکنش Settle (تسویه)نشده است',
    '47': 'تراکنش Settle یافت نشد',
    '48': 'تراکنش Reverse شده است',
    '49': 'تراکنش Refund یافت نشد',
    '51': 'تراکنش تکراری است',
    '54': 'تراکنش مرجع موجود نیست',
    '55': 'تراکنش نامعتبر است',
    '61': 'خطا در واریز',
    '111': 'صادر کننده کارت نامعتبر است',
    '112': 'خطای سوییچ صادر کننده کارت',
    '113': 'پاسخی از صادر کننده کارت دریافت نشد',
    '114': 'دارنده کارت مجاز به انجام این تراکنش نیست',
    '412': 'شناسه قبض نادرست است',
    '413': 'شناسه پرداخت نادرست است',
    '414': 'سازمان صادر کننده قبض نامعتبر است',
    '415': 'زمان جلسه کاری به پایان رسیده است',
    '416': 'خطا در ثبت اطلاعات',
    '417': 'شناسه پرداخت کننده نامعتبر است',
    '418': 'اشکال در تعریف اطلاعات مشتری',
    '419': 'تعداد دفعات ورود اطلاعات از حد مجاز گذشته است',
    '421': 'IP نامعتبر است',
}

def payment_page(request):
    """
    Initiates a payment request with the Beh Pardakht gateway.
    """
    pending_appointment_id = request.session.get('pending_appointment_id')
    if not pending_appointment_id:
        return redirect('booking:doctor_list')

    # 🟢 خطوط ۱۸-۲۲: تبدیل امن شناسه سفارش به عدد صحیح
    try:
        order_id_int = int(pending_appointment_id)
    except (ValueError, TypeError):
        error_message = "خطا: شناسه سفارش برای ارسال به بانک نامعتبر است."
        return render(request, 'booking/payment_page.html', {'error_message': error_message})

    appointment = get_object_or_404(Appointment, pk=order_id_int) # استفاده از order_id_int
    
    verified_phone = request.session.pop('verified_patient_phone', None)

    # Generate a unique order ID for this specific payment attempt
    unique_order_id = int(f"{appointment.id}{int(time.time())}")
    appointment.payment_order_id = unique_order_id
    appointment.save()

    from zeep import Client
    
    client = Client('https://bpm.shaparak.ir/pgwchannel/services/pgw?wsdl')
    
    terminal_id = settings.BEH_PARDAKHT_TERMINAL_ID
    user_name = settings.BEH_PARDAKHT_USERNAME
    user_password = settings.BEH_PARDAKHT_PASSWORD
    order_id = unique_order_id # ⬅️ استفاده از متغیر ایمن شده
    amount = int(appointment.doctor.visit_fee)
    local_date = datetime.datetime.now().strftime('%Y%m%d')
    local_time = datetime.datetime.now().strftime('%H%M%S')
    additional_data = f'Appointment for {appointment.patient_name}'
    callback_url = request.build_absolute_uri(reverse('booking:verify_payment')).replace("http://", "https://")
    payer_id = 0
    
    try:
        result = client.service.bpPayRequest(
            terminalId=terminal_id,
            userName=user_name,
            userPassword=user_password,
            orderId=order_id,
            amount=amount,
            localDate=local_date,
            localTime=local_time,
            additionalData=additional_data,
            callBackUrl=callback_url,
            payerId=payer_id
        )
        
        # ⭐️ خطوط ۴۰-۵۸: اصلاح حیاتی برای مدیریت خطای unpack
        if ',' in result:
            res_code, ref_id = result.split(',')
            if res_code == '0' and ref_id:
                context = {
                    'ref_id': ref_id,
                    'post_url': 'https://bpm.shaparak.ir/pgwchannel/startpay.mellat',
                    'appointment': appointment,
                    'payment_amount': amount,
                    'page_title': 'صفحه پرداخت',
                    'error_message': None,
                    'verified_phone': verified_phone
                }
                return render(request, 'booking/payment_page.html', context)
            else:
                error_message = MELLAT_BANK_ERRORS.get(res_code, f"خطای نامشخص از بانک: {res_code}")
        else:
            # پاسخ فقط کد خطا است
            res_code = result
            error_message = MELLAT_BANK_ERRORS.get(res_code, f"خطای نامشخص از بانک: {res_code}")
            
    except Exception as e:
        error_message = f"خطا در برقراری ارتباط با درگاه پرداخت: {e}"

    context = {
        'appointment': appointment,
        'payment_amount': amount,
        'page_title': 'صفحه پرداخت',
        'error_message': error_message
    }
    return render(request, 'booking/payment_page.html', context)

@csrf_exempt
def verify_payment(request):
    res_code = request.POST.get('ResCode')
    # 🟢 خطوط ۷-۸: تغییر نام متغیرهای دریافتی برای تمایز با نسخه عددی
    sale_order_id_str = request.POST.get('SaleOrderId') or request.POST.get('saleOrderId')
    sale_reference_id_str = request.POST.get('SaleReferenceId') or request.POST.get('saleReferenceId')

    payment_successful = False
    message = ''

    # 1. Check if the initial transaction was successful at the bank's end.
    if res_code != '0':
        message = MELLAT_BANK_ERRORS.get(res_code, f"تراکنش ناموفق بود. کد خطا: {res_code}")
        return render(request, 'booking/payment_result.html', {
            'payment_successful': False, 'message': message, 'page_title': 'نتیجه پرداخت'
        })

    # 2. If successful, proceed to verify and settle.
    # 🟢 خطوط ۲۶-۲۹: تبدیل امن به عدد صحیح
    try:
        sale_order_id_int = int(sale_order_id_str)
        sale_reference_id_int = int(sale_reference_id_str)
    except (ValueError, TypeError):
        message = "خطا: شناسه تراکنش نامعتبر است."
        return render(request, 'booking/payment_result.html', {
            'payment_successful': False, 'message': message, 'page_title': 'نتیجه پرداخت'
        })

    try:
        from zeep import Client
        client = Client('https://bpm.shaparak.ir/pgwchannel/services/pgw?wsdl')

        terminal_id = settings.BEH_PARDAKHT_TERMINAL_ID
        user_name = settings.BEH_PARDAKHT_USERNAME
        user_password = settings.BEH_PARDAKHT_PASSWORD
        
        # ⭐️ خطوط ۳۹-۴۱: استفاده از نسخه‌های عددی برای متد zeep
        common_params = {
            'terminalId': terminal_id, 'userName': user_name, 'userPassword': user_password,
            'orderId': sale_order_id_int, 'saleOrderId': sale_order_id_int, 'saleReferenceId': sale_reference_id_int
        }

        verify_result = str(client.service.bpVerifyRequest(**common_params))

        if verify_result == '0':
            # 3. Payment is verified, now settle it.
            settle_result = str(client.service.bpSettleRequest(**common_params))
            if settle_result == '0':
                # 4. All steps successful. Finalize appointment.
                # ⭐️ خط ۴۹: استفاده از نسخه عددی برای کوئری دیتابیس
                appointment = get_object_or_404(Appointment, payment_order_id=sale_order_id_int)
                appointment.status = 1
                appointment.save()

                # --- Send SMS Confirmation ---
                try:
                    # ✜ بخش ارسال پیامک تایید نوبت ✜

                    tehran_tz = pytz.timezone('Asia/Tehran')
                    appointment_datetime_local = appointment.appointment_datetime.astimezone(tehran_tz)
                    jalali_datetime = jdatetime.datetime.fromgregorian(datetime=appointment_datetime_local)
                    formatted_time = jalali_datetime.strftime('%Y/%m/%d ساعت %H:%M')
                    bimar = appointment.patient_name
                    dr = appointment.doctor.user.get_full_name()
                    time = formatted_time
                    adders = appointment.doctor.address
                    tel = appointment.doctor.phone_number
                    pattern_values = f"{bimar},{dr},{time},{adders},{tel}"
                    AMOOT_SMS_API_TOKEN = settings.AMOOT_SMS_API_TOKEN
                    AMOOT_SMS_API_URL = settings.AMOOT_SMS_API_URL
                    payload = {
                        'token': AMOOT_SMS_API_TOKEN,
                        'Mobile': appointment.patient_phone,
                        'PatternCodeID': 4161,
                        'PatternValues': pattern_values,
                    }
                    response = requests.post(AMOOT_SMS_API_URL, data=payload)
                    logging.info(f"SMS API Response for order {sale_order_id_int}: {response.status_code} - {response.text}")

                except requests.exceptions.RequestException as e:
                    # حتی اگر پیامک ارسال نشود، نباید جلوی تکمیل فرآیند نوبت‌گیری را بگیرد
                    logging.error(f"خطا در ارسال پیامک تایید نوبت برای سفارش {sale_order_id_int}: {e}")
                except Exception as e:
                    logging.error(f"یک خطای پیش‌بینی نشده در ارسال پیامک برای سفارش {sale_order_id_int} رخ داد: {e}")


                login(request, appointment.patient)
                request.session.save()
                messages.success(request, "پرداخت با موفقیت انجام شد و نوبت شما ثبت گردید.")
                return redirect('booking:patient_dashboard')
            else:
                # 5. Settle failed, reverse the transaction.
                message = f"خطا در تسویه حساب: {MELLAT_BANK_ERRORS.get(settle_result, settle_result)}"
                reversal_result = str(client.service.bpReversalRequest(**common_params))
                if reversal_result == '0':
                    message += " (مبلغ با موفقیت به حساب شما بازگردانده شد)."
                else:
                    message += f" (خطا در بازگشت وجه: {MELLAT_BANK_ERRORS.get(reversal_result, reversal_result)})."
        else:
            # 6. Verify failed, reverse the transaction.
            message = f"خطا در تایید پرداخت: {MELLAT_BANK_ERRORS.get(verify_result, verify_result)}"
            reversal_result = str(client.service.bpReversalRequest(**common_params))
            if reversal_result == '0':
                message += " (مبلغ با موفقیت به حساب شما بازگردانده شد)."
            else:
                message += f" (خطا در بازگشت وجه: {MELLAT_BANK_ERRORS.get(reversal_result, reversal_result)})."

    except Exception as e:
        message = f"خطا در ارتباط با وب سرویس به پرداخت: {e}. لطفاً با پشتیبانی تماس بگیرید."

    return render(request, 'booking/payment_result.html', {
        'payment_successful': payment_successful, 'message': message, 'page_title': 'نتیجه پرداخت'
    })


def initiate_payment(request, appointment_id):
    """
    شروع فرآیند پرداخت و هدایت به درگاه
    """
    appointment = get_object_or_404(Appointment, pk=appointment_id)
    
    # ذخیره appointment_id در session
    request.session['pending_appointment_id'] = appointment_id
    
    return redirect('booking:payment_page')

def confirm_payment(request):
    """
    پردازش پرداخت شبیه‌سازی شده و نهایی کردن نوبت با استفاده از session.
    """
    pending_appointment_id = request.session.get('pending_appointment_id')
    if not pending_appointment_id:
        return redirect('booking:doctor_list')

    appointment = get_object_or_404(Appointment, pk=pending_appointment_id)

    appointment.status = 1
    appointment.save()

    # Clear the session variable after successful booking
    del request.session['pending_appointment_id']

    context = {
        'appointment': appointment,
        'page_title': 'نوبت شما با موفقیت ثبت شد'
    }
    return render(request, 'booking/confirmation_page.html', context)

from django.contrib.auth import login, logout
from django.http import JsonResponse
from django.forms import modelformset_factory
from .forms import AppointmentUpdateForm, DailyExpenseForm, DoctorRegistrationForm, UserUpdateForm, DoctorProfileUpdateForm, SecretarySignUpForm
from .models import DailyExpense
from django.db.models import Count

@login_required
def secretary_panel(request, date=None):
    """
    پنل مدیریت منشی (داشبورد).
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    current_date = datetime.date.today()
    if date:
        try:
            current_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            pass


    end_date = current_date + datetime.timedelta(days=45)
    availabilities = doctor_profile.availabilities.filter(is_active=True)

    # Optimize appointment counting
    booked_appointments = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date__range=[current_date, end_date],
        status__in=[1, 2, 4]
    ).values('appointment_datetime__date').annotate(count=Count('id'))

    booked_counts = {item['appointment_datetime__date']: item['count'] for item in booked_appointments}

    # Get future available days for manual booking
    future_days_info = []
    for i in range(0, 46):  # From today for the next 45 days
        future_date = current_date + datetime.timedelta(days=i)
        daily_availabilities = availabilities.filter(day_of_week=future_date.weekday())

        if daily_availabilities.exists():
            total_capacity = sum(da.visit_count for da in daily_availabilities)
            day_info = {'date': future_date, 'booked_percentage': 0}

            if total_capacity > 0:
                booked_count = booked_counts.get(future_date, 0)
                booked_percentage = (booked_count / total_capacity) * 100
                day_info['booked_percentage'] = booked_percentage

            future_days_info.append(day_info)

    context = {
        'today': current_date,
        'future_days': future_days_info,
        'page_title': 'تقویم نوبت‌دهی دستی توسط منشی'
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'booking/secretary_panel_content.html', context)
    return render(request, 'booking/secretary_panel.html', context)


from .models import InsuranceFee

@login_required
def patient_list(request):
    """
    نمایش لیست تمام بیماران با قابلیت جستجو.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    queryset = Appointment.objects.filter(
        doctor=doctor_profile, status=1
    ).order_by('-appointment_datetime')

    query = request.GET.get('q')
    if query:
        queryset = queryset.filter(
            Q(patient_name__icontains=query) |
            Q(patient_national_id__icontains=query) |
            Q(patient_phone__icontains=query) |
            Q(service_description__icontains=query)
        )

    context = {
        'appointments': queryset,
        'page_title': 'لیست تمام بیماران',
        'search_query': query or ''
    }
    return render(request, 'booking/patient_list.html', context)


@login_required
def daily_patients(request, date=None):
    """
    نمایش و مدیریت لیست بیماران امروز.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    if date:
        try:
            current_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            current_date = datetime.date.today()
    else:
        current_date = datetime.date.today()

    AppointmentFormSet = modelformset_factory(Appointment, form=AppointmentUpdateForm, extra=0)

    queryset = Appointment.objects.filter(
        doctor=doctor_profile, appointment_datetime__date=current_date,   status__in=[1, 2]
    ).order_by('appointment_datetime')

    if request.method == 'POST':
        formset = AppointmentFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            appointments = formset.save(commit=False)
            for appointment in appointments:
                # Update status based on payment method
                if appointment.payment_method and appointment.payment_method >= 1:
                    appointment.status = 2
                else:
                    appointment.status = 1
                appointment.save()

                # Update insurance fee if visit fee was paid
                if appointment.visit_fee_paid is not None:
                    InsuranceFee.objects.update_or_create(
                        doctor=doctor_profile,
                        insurance_type=appointment.insurance_type,
                        defaults={'fee': appointment.visit_fee_paid}
                    )

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            return redirect('booking:daily_patients', date=date)
        elif request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # Handle formset errors for AJAX requests
            return JsonResponse({'success': False, 'errors': formset.errors})
    else:
        # Pre-fill visit fee based on insurance
        insurance_fees = {
            fee.insurance_type: fee.fee
            for fee in InsuranceFee.objects.filter(doctor=doctor_profile)
        }
        for appointment in queryset:
            if not appointment.visit_fee_paid:
                appointment.visit_fee_paid = insurance_fees.get(appointment.insurance_type)

        formset = AppointmentFormSet(queryset=queryset)

    context = {
        'formset': formset,
        'today': current_date,
        'page_title': 'لیست بیماران امروز'
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'booking/daily_patients_content.html', context)
    return render(request, 'booking/daily_patients.html', context)


@login_required
def secretary_payments(request, date=None):
    """
    نمایش و ثبت هزینه‌های روزانه منشی.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    if date:
        try:
            current_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            current_date = datetime.date.today()
    else:
        current_date = datetime.date.today()


    if request.method == 'POST':
        expense_form = DailyExpenseForm(request.POST)
        if expense_form.is_valid():
            expense = expense_form.save(commit=False)
            expense.doctor = doctor_profile
            expense.date = current_date
            expense.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'description': expense.description,
                    'amount': expense.amount
                })
            return redirect('booking:secretary_payments', date=date)
        elif request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': expense_form.errors})

    expense_form = DailyExpenseForm()
    daily_expenses = DailyExpense.objects.filter(doctor=doctor_profile, date=current_date)

    # Calculate previous day's balance
    yesterday = current_date - datetime.timedelta(days=1)
    previous_cash_income = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date__lte=yesterday,
        payment_method=2,  # نقدی
        visit_fee_paid__isnull=False
    ).aggregate(total=Sum('visit_fee_paid'))['total'] or 0
    previous_expenses_and_payments = DailyExpense.objects.filter(
        doctor=doctor_profile,
        date__lte=yesterday
    ).aggregate(total=Sum('amount'))['total'] or 0
    previous_day_balance = previous_cash_income - previous_expenses_and_payments

  # Calculate today's cash income
    todays_cash_income = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date=current_date,
        payment_method=2,  # نقدی
        visit_fee_paid__isnull=False
    ).aggregate(total=Sum('visit_fee_paid'))['total'] or 0

    # Calculate current secretary cash box balance
    total_cash_income_lte = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date__lte=current_date,
        payment_method=2,  # نقدی
        visit_fee_paid__isnull=False
    ).aggregate(total=Sum('visit_fee_paid'))['total'] or 0

    total_expenses_and_payments = DailyExpense.objects.filter(
        doctor=doctor_profile,
        date__lte=current_date
    ).aggregate(total=Sum('amount'))['total'] or 0

    cash_box_balance = total_cash_income_lte - total_expenses_and_payments

    context = {
        'expense_form': expense_form,
        'daily_expenses': daily_expenses,
        'today': current_date,
        'page_title': 'پرداخت‌های منشی',
        'cash_box_balance': cash_box_balance,
        'previous_day_balance': previous_day_balance,
        'todays_cash_income':todays_cash_income,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'booking/secretary_payments_content.html', context)
    return render(request, 'booking/secretary_payments.html', context)

@login_required
def manage_day(request, date):
    """
    صفحه مدیریت روزانه برای منشی، شامل ثبت نوبت، مسدود کردن و فعال‌سازی نوبت‌ها.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    try:
        # The date comes from the URL in Jalali format. Replace '/' with '-' to support both formats.
        normalized_date = date.replace('/', '-')
        jalali_date = jdatetime.datetime.strptime(normalized_date, '%Y-%m-%d').date()
        target_date = jalali_date.togregorian()
    except ValueError:
        return redirect('booking:secretary_panel')

    jalali_day_names = ["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنجشنبه", "جمعه"]
    persian_weekday = jalali_day_names[jalali_date.weekday()]

    # --- Logic to calculate and display all time slots ---
    all_slots = []
    booked_datetimes = list(Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date=target_date,
        status__in=[1, 2, 4]
    ).values_list('appointment_datetime', flat=True))

    time_slot_exceptions = TimeSlotException.objects.filter(
        doctor=doctor_profile,
        datetime_slot__date=target_date
    )

    canceled_slots = list(time_slot_exceptions.filter(is_cancellation=True).values_list('datetime_slot', flat=True))
    added_slots = list(time_slot_exceptions.filter(is_cancellation=False).values_list('datetime_slot', flat=True))

    availabilities = DoctorAvailability.objects.filter(
        doctor=doctor_profile,
        day_of_week=target_date.weekday(),
        is_active=True
    )

    for avail in availabilities:
        duration = (datetime.datetime.combine(target_date, avail.end_time) - datetime.datetime.combine(target_date, avail.start_time))
        if avail.visit_count > 0:
            interval = duration / avail.visit_count
            current_time_naive = datetime.datetime.combine(target_date, avail.start_time)
            for i in range(avail.visit_count):
                current_time_aware = timezone.make_aware(current_time_naive)
                status = 'available'
                if current_time_aware in booked_datetimes:
                    status = 'booked'
                elif current_time_aware in canceled_slots:
                    status = 'canceled'

                all_slots.append({'time': current_time_aware, 'status': status})
                current_time_naive += interval

    for added_slot in added_slots:
        if added_slot not in booked_datetimes and added_slot not in canceled_slots:
            all_slots.append({'time': added_slot, 'status': 'available'})

    # Handle POST requests for booking, blocking, or unblocking slots
    if request.method == 'POST':
        action = request.POST.get('action')
        slot_iso = request.POST.get('selected_slot')

        if action and slot_iso:
            slot_datetime = datetime.datetime.fromisoformat(slot_iso)

            if action == 'block':
                TimeSlotException.objects.get_or_create(
                    doctor=doctor_profile,
                    datetime_slot=slot_datetime
                )
                return redirect('booking:manage_day', date=date)

            elif action == 'unblock':
                TimeSlotException.objects.filter(
                    doctor=doctor_profile,
                    datetime_slot=slot_datetime
                ).delete()
                return redirect('booking:manage_day', date=date)

            elif action == 'book':
                form = AppointmentBookingForm(request.POST)
                if form.is_valid():
                    with transaction.atomic():
                        # Find or create a patient user
                        patient_user, created = User.objects.get_or_create(
                            username=form.cleaned_data['patient_phone'],
                            defaults={
                                'first_name': form.cleaned_data['patient_name'],
                                'user_type': 'PATIENT'
                            }
                        )
                        # Create the appointment
                        appointment = form.save(commit=False)
                        appointment.doctor = doctor_profile
                        appointment.patient = patient_user
                        appointment.appointment_datetime = slot_datetime
                        appointment.status = 1
                        appointment.save()
                        return redirect('booking:manage_day', date=date)
                # If form is invalid, we will fall through and re-render the page with errors

        elif action == 'add_slot':
            last_slot_time = datetime.time(8, 50) # Default start time if no slots exist
            if all_slots:
                last_slot_time = max(s['time'] for s in all_slots).time()

            new_slot_datetime_naive = datetime.datetime.combine(target_date, last_slot_time) + datetime.timedelta(minutes=10)
            new_slot_datetime_aware = timezone.make_aware(new_slot_datetime_naive)

            TimeSlotException.objects.create(
                doctor=doctor_profile,
                datetime_slot=new_slot_datetime_aware,
                is_cancellation=False # This is an addition, not a cancellation
            )
            return redirect('booking:manage_day', date=date)

    # If the request was a failed POST for booking, use that form, otherwise create a new one
    if request.method == 'POST' and 'form' in locals():
        booking_form = form
    else:
        booking_form = AppointmentBookingForm()

    context = {
        'doctor': doctor_profile,
        'date': target_date,
        'jalali_date_str': date,
        'all_slots': sorted(all_slots, key=lambda x: x['time']),
        'form': booking_form,
        'has_availability': bool(availabilities),
        'page_title': f'مدیریت نوبت‌های روز {jalali_date.strftime("%A")} {jalali_date.strftime("%Y/%m/%d")}'    
    }
    return render(request, 'booking/manage_day.html', context)


def doctor_signup(request):
    if request.method == 'POST':
        form = DoctorRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # Deactivate account until verification
            user.save()

            # Create the doctor's profile
            DoctorProfile.objects.create(
                user=user,
                specialty=form.cleaned_data.get('specialty'),
                address=form.cleaned_data.get('address'),
                phone_number=form.cleaned_data.get('phone_number'),
                mobile_number=form.cleaned_data.get('mobile_number'),
                photo=form.cleaned_data.get('photo'),
                biography=form.cleaned_data.get('biography')
            )

            try:
                AMOOT_SMS_API_TOKEN=settings.AMOOT_SMS_API_TOKEN
                AMOOT_SMS_API_URL=settings.AMOOT_SMS_API_URL
                otp_code = str(random.randint(100000, 999999))
                request.session['otp_code'] = otp_code
                request.session['new_user_id'] = user.id
                mob_number = form.cleaned_data.get('mobile_number')

                payload = {
                       'token': AMOOT_SMS_API_TOKEN,
                       'Mobile': mob_number,
                       'PatternCodeID':4018,
                       'PatternValues': otp_code,
                    }
                requests.post(AMOOT_SMS_API_URL,data=payload)
                return redirect('booking:verify_doctor_signup')
            except requests.exceptions.RequestException as e:
                # Handle exceptions, delete the created user since we can't verify them
                user.delete()
                form.add_error(None, "خطا در ارسال کد تایید. لطفاً مجدداً تلاش کنید.")
                # Fall through to render the form with the error message
    else: # GET
        form = DoctorRegistrationForm()

    return render(request, 'booking/signup.html', {'form': form, 'page_title': 'ثبت نام پزشک'})

def verify_doctor_signup(request):
    new_user_id = request.session.get('new_user_id')
    if not new_user_id:
        return redirect('signup')

    try:
        user = User.objects.get(pk=new_user_id)
    except User.DoesNotExist:
        return redirect('signup')

    if request.method == 'POST':
        otp_from_user = request.POST.get('otp')
        otp_from_session = request.session.get('otp_code')

        if otp_from_user == otp_from_session:
            user.is_active = True
            user.save()

            # Clean up session
            del request.session['new_user_id']
            del request.session['otp_code']

            login(request, user)
            return redirect('booking:doctor_dashboard')
        else:
            error = 'کد وارد شده صحیح نمی‌باشد.'
            return render(request, 'booking/verify_doctor_signup.html', {'error': error})

    return render(request, 'booking/verify_doctor_signup.html', {'page_title': 'تأیید ثبت نام پزشک'})

@login_required
@doctor_required
def edit_profile(request):
    doctor_profile = request.user.doctor_profile

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = DoctorProfileUpdateForm(request.POST, request.FILES, instance=doctor_profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('booking:edit_profile') # Redirect back to the same page to show success
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = DoctorProfileUpdateForm(instance=doctor_profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'page_title': 'ویرایش پروفایل'
    }
    return render(request, 'booking/edit_profile.html', context)

from django.db.models import Sum, Q, Count, Avg

# ... (other code)

@login_required
def edit_expense(request, pk):
    """
    ویرایش یک هزینه یا پرداخت ثبت شده.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    expense = get_object_or_404(DailyExpense, pk=pk, doctor=doctor_profile)

    if request.user.user_type == 'SECRETARY' and expense.date != datetime.date.today():
        return redirect('booking:secretary_payments')

    if request.method == 'POST':
        form = DailyExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            # Redirect to the secretary_payments page for the date of the expense
            return redirect(reverse('booking:secretary_payments', kwargs={'date': expense.date.strftime('%Y-%m-%d')}))
    else:
        form = DailyExpenseForm(instance=expense)

    context = {
        'form': form,
        'page_title': 'ویرایش هزینه/پرداخت',
        'expense_date': expense.date
    }
    return render(request, 'booking/edit_expense.html', context)

@login_required
def delete_expense(request, pk):
    """
    حذف یک هزینه یا پرداخت ثبت شده.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    expense = get_object_or_404(DailyExpense, pk=pk, doctor=doctor_profile)

    if request.user.user_type == 'SECRETARY' and expense.date != datetime.date.today():
        return redirect('booking:secretary_payments')

    expense_date_str = expense.date.strftime('%Y-%m-%d')

    if request.method == 'POST':
        expense.delete()
        return redirect(reverse('booking:secretary_payments', kwargs={'date': expense_date_str}))

    context = {
        'expense': expense,
        'page_title': 'تایید حذف',
        'cancel_url': reverse('booking:secretary_payments', kwargs={'date': expense_date_str})
    }
    return render(request, 'booking/delete_expense_confirm.html', context)


from .decorators import doctor_required

@login_required
def financial_report(request, period='daily', date=None):
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    if request.user.user_type == 'SECRETARY' and period in ['monthly', 'yearly']:
        return redirect('booking:doctor_dashboard')

    # Determine the date range based on the period
    if date:
        try:
            current_date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            current_date = datetime.date.today()
    else:
        current_date = datetime.date.today()

    jalali_today = jdatetime.date.fromgregorian(date=current_date)
    end_date = current_date
    page_title = 'گزارش مالی'

    if period == 'daily':
        start_date = current_date
        page_title = 'گزارش مالی روزانه'
    elif period == 'monthly':
        start_date = jdatetime.date(jalali_today.year, jalali_today.month, 1).togregorian()
        page_title = f'گزارش مالی ماهانه ({jalali_today.strftime("%B %Y")})'
    elif period == 'yearly':
        start_date = jdatetime.date(jalali_today.year, 1, 1).togregorian()
        page_title = f'گزارش مالی سالانه ({jalali_today.strftime("%Y")})'
    else: # Default to daily if period is invalid
        period = 'daily'
        start_date = current_date
        page_title = 'گزارش مالی روزانه'


    # --- Calculations for the selected period ---
    all_appointments_in_period = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date__range=[start_date, end_date]
    )
    appointments_in_period = all_appointments_in_period.filter(
        visit_fee_paid__isnull=False
    )
    expenses_in_period_queryset = DailyExpense.objects.filter(
        doctor=doctor_profile,
        date__range=[start_date, end_date]
    )
    total_income = appointments_in_period.aggregate(total=Sum('visit_fee_paid'))['total'] or 0
    income_by_payment_method_values = appointments_in_period.values('payment_method').annotate(total=Sum('visit_fee_paid'))
    income_by_payment_method_dict = {
        dict(Appointment.PAYMENT_METHOD_CHOICES).get(item['payment_method'], 'نامشخص'): item['total']
        for item in income_by_payment_method_values
    }
    income_by_insurance_values = appointments_in_period.values('insurance_type').annotate(total=Sum('visit_fee_paid'), count=Count('id'))
    income_by_insurance_data = {
        dict(Appointment.INSURANCE_CHOICES).get(item['insurance_type'], 'نامشخص'): {
            'total': item['total'],
            'count': item['count']
        }
        for item in income_by_insurance_values
    }
    total_period_entries = expenses_in_period_queryset.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = abs(sum(item.amount for item in expenses_in_period_queryset if item.amount > 0))
    total_payments_received = sum(item.amount for item in expenses_in_period_queryset if item.amount < 0)
    net_income = total_income - total_period_entries

    # --- Secretary Cash Box Calculation (Cumulative up to end_date) ---
    # This calculation should always be cumulative up to the selected date.
    total_cash_income = Appointment.objects.filter(
        doctor=doctor_profile,
        appointment_datetime__date__lte=end_date,
        payment_method=2,  # نقدی
        visit_fee_paid__isnull=False
    ).aggregate(total=Sum('visit_fee_paid'))['total'] or 0

    total_expenses_and_payments = DailyExpense.objects.filter(
        doctor=doctor_profile,
        date__lte=end_date
    ).aggregate(total=Sum('amount'))['total'] or 0

    cash_box_balance = total_cash_income - total_expenses_and_payments

    if request.method == 'POST' and 'settle_up' in request.POST:
        if cash_box_balance != 0:
            DailyExpense.objects.create(
                doctor=doctor_profile,
                date=current_date, # Settle on the current day
                description="تسویه صندوق منشی",
                amount=cash_box_balance
            )
            # Redirect to prevent form resubmission
            return redirect('booking:financial_report', period=period, date=current_date.strftime('%Y-%m-%d'))

    total_booked_count = all_appointments_in_period.exclude(status=3).count()
    total_visited_count = appointments_in_period.count()

    context = {
        'today': current_date, # 'today' is used for navigation, so keep it
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'page_title': page_title,
        'total_income': total_income,
        'income_by_payment_method': income_by_payment_method_dict,
        'income_by_insurance': income_by_insurance_data,
        'todays_expenses_queryset': expenses_in_period_queryset, # Renaming this might be good, but let's keep it for now to avoid breaking the template
        'total_expenses': total_expenses,
        'total_payments_received': total_payments_received,
        'net_income': net_income,
        'cash_box_balance': cash_box_balance,
        'total_booked_count': total_booked_count,
        'total_visited_count': total_visited_count,
    }

    if period in ['monthly', 'yearly']:
        grouped_expenses = expenses_in_period_queryset.filter(amount__gt=0).values('description').annotate(
            total=Sum('amount'),
            count=Count('id'),
            average=Avg('amount')
        ).order_by('-total')
        grouped_payments = expenses_in_period_queryset.filter(amount__lt=0).values('description').annotate(
            total=Sum('amount'),
            count=Count('id'),
            average=Avg('amount')
        ).order_by('total')
        context['grouped_expenses'] = grouped_expenses
        context['grouped_payments'] = grouped_payments

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'booking/financial_report_content.html', context)
    return render(request, 'booking/financial_report.html', context)

@login_required
@doctor_required
def expense_balance_report(request):
    """
    گزارش تراز هزینه سالانه.
    """
    if request.user.user_type == 'DOCTOR':
        doctor_profile = request.user.doctor_profile
    elif request.user.user_type == 'SECRETARY':
        doctor_profile = request.user.doctor
    else:
        return redirect('booking:doctor_list')
    end_date = datetime.date.today()
    jalali_today = jdatetime.date.fromgregorian(date=end_date)
    start_of_year = jdatetime.date(jalali_today.year, 1, 1).togregorian()
    start_date = start_of_year

    expenses = DailyExpense.objects.filter(
        doctor=doctor_profile,
        date__range=[start_date, end_date],
        amount__gt=0  # Only include expenses, not payments received
    ).values('description').annotate(
        count=Count('id'),
        total_amount=Sum('amount'),
        average_amount=Avg('amount')
    ).order_by('-total_amount')

    total_expense_sum = sum(item['total_amount'] for item in expenses)

    context = {
        'expenses': expenses,
        'start_date': start_date,
        'end_date': end_date,
        'total_expense_sum': total_expense_sum,
        'page_title': f'خلاصه صورت هزینه های مطب از تاریخ {jdatetime.date.fromgregorian(date=start_date).strftime("%Y/%m/%d")} تا تاریخ {jdatetime.date.fromgregorian(date=end_date).strftime("%Y/%m/%d")}'
    }

    return render(request, 'booking/expense_balance_report.html', context)

@login_required
def expense_item_details(request, description):
    """
    نمایش لیست جزئیات یک هزینه خاص.
    """
    if not request.user.user_type == 'DOCTOR':
        return redirect('booking:doctor_list')

    doctor_profile = request.user.doctor_profile
    end_date = datetime.date.today()
    jalali_today = jdatetime.date.fromgregorian(date=end_date)
    start_of_year = jdatetime.date(jalali_today.year, 1, 1).togregorian()
    start_date = start_of_year

    expenses = DailyExpense.objects.filter(
        doctor=doctor_profile,
        description=description,
        date__range=[start_date, end_date],
        amount__gt=0
    ).order_by('date')

    context = {
        'expenses': expenses,
        'description': description,
        'start_date': start_date,
        'end_date': end_date,
        'page_title': f'لیست هزینه ها ({description}) از تاریخ {jdatetime.date.fromgregorian(date=start_date).strftime("%Y/%m/%d")} تا تاریخ {jdatetime.date.fromgregorian(date=end_date).strftime("%Y/%m/%d")}'
    }

    return render(request, 'booking/expense_item_details.html', context)


def accounting_guide(request):
    """
    Displays the accounting guide page.
    """
    return render(request, 'booking/accounting_guide.html')


def secretary_signup(request):
    if request.method == 'POST':
        form = SecretarySignUpForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # Deactivate account until verification
            user.save()

            try:
                AMOOT_SMS_API_TOKEN=settings.AMOOT_SMS_API_TOKEN
                AMOOT_SMS_API_URL=settings.AMOOT_SMS_API_URL
                otp_code = str(random.randint(100000, 999999))
                request.session['otp_code'] = otp_code
                request.session['new_user_id'] = user.id
                doctor_mobile_number = user.doctor.mobile_number

                payload = {
                       'token': AMOOT_SMS_API_TOKEN,
                       'Mobile': doctor_mobile_number,
                       'PatternCodeID':4018, # Assuming this is a generic OTP pattern
                       'PatternValues': otp_code,
                    }
                requests.post(AMOOT_SMS_API_URL,data=payload)
                return redirect('booking:verify_secretary_signup')
            except requests.exceptions.RequestException as e:
                user.delete()
                form.add_error(None, "خطا در ارسال کد تایید. لطفاً مجدداً تلاش کنید.")
    else:
        form = SecretarySignUpForm()

    return render(request, 'booking/secretary_signup.html', {'form': form, 'page_title': 'ثبت نام منشی'})


def verify_secretary_signup(request):
    new_user_id = request.session.get('new_user_id')
    if not new_user_id:
        return redirect('booking:secretary_signup')

    try:
        user = User.objects.get(pk=new_user_id)
    except User.DoesNotExist:
        return redirect('booking:secretary_signup')

    if request.method == 'POST':
        otp_from_user = request.POST.get('otp')
        otp_from_session = request.session.get('otp_code')

        if otp_from_user == otp_from_session:
            user.is_active = True
            user.save()

            del request.session['new_user_id']
            del request.session['otp_code']

            login(request, user)
            return redirect('booking:doctor_dashboard')
        else:
            error = 'کد وارد شده صحیح نمی‌باشد.'
            return render(request, 'booking/verify_secretary_signup.html', {'error': error})

    return render(request, 'booking/verify_secretary_signup.html', {'page_title': 'تأیید ثبت نام منشی'})


@login_required
def export_patients_to_excel(request):
    """
    Export patient list to Excel.
    """
    if not request.user.user_type == 'DOCTOR':
        return redirect('booking:doctor_list')

    doctor_profile = request.user.doctor_profile
    queryset = Appointment.objects.filter(
        doctor=doctor_profile, status=1
    ).order_by('-appointment_datetime')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=patients.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Patients'

    columns = [
        'نام بیمار',
        'کد ملی',
        'شماره همراه',
        'نوع بیمه',
        'زمان نوبت',
        'شرح خدمات',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for appointment in queryset:
        row_num += 1
        row = [
            appointment.patient_name,
            appointment.patient_national_id,
            appointment.patient_phone,
            appointment.get_insurance_type_display(),
            appointment.appointment_datetime.strftime('%Y-%m-%d %H:%M'),
            appointment.service_description,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


from itertools import groupby
from operator import attrgetter

@login_required
def reservation_list(request):
    """
    Displays a list of all future reserved appointments for the doctor/secretary.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    today = datetime.date.today()
    reservations_qs = Appointment.objects.filter(
        doctor=doctor_profile,
        status=1,
        appointment_datetime__gte=today
    ).order_by('appointment_datetime')

    colors = ["#E0FFFF", "#FFFACD", "#FFE4E1", "#F0FFF0", "#F0F8FF", "#E6E6FA", "#FAFAD2"]
    grouped_reservations = []

    # Pre-process to add a 'date' attribute for grouping
    reservations_list = []
    for r in reservations_qs:
        r.date = r.appointment_datetime.date()
        reservations_list.append(r)

    for i, (date, group) in enumerate(groupby(reservations_list, key=attrgetter('date'))):
        group_list = list(group)
        color = colors[i % len(colors)]
        grouped_reservations.append({
            'date': date,
            'reservations': group_list,
            'color': color
        })

    context = {
        'grouped_reservations': grouped_reservations,
        'page_title': 'لیست رزروها'
    }
    return render(request, 'booking/reservation_list.html', context)


from django.views.decorators.http import require_POST

@login_required
@require_POST
def cancel_reservation(request, pk):
    """
    Cancels a reservation.
    """
    doctor_profile = _get_doctor_profile(request.user)
    if not doctor_profile:
        return redirect('booking:doctor_list')

    reservation = get_object_or_404(Appointment, pk=pk, doctor=doctor_profile)
    reservation.delete()

    return redirect('booking:reservation_list')


from django.contrib.auth import login, logout

def patient_dashboard_entry(request):
    """
    Entry point for patient dashboard.
    This page will have JS to check localStorage and redirect.
    """
    return render(request, 'booking/patient_dashboard_entry.html')

def patient_login(request):
    """
    Handles patient login by sending an OTP to their mobile number.
    """
    if request.method == 'POST':
        mobile_number = request.POST.get('mobile_number', '').strip()
        if not (mobile_number.isdigit() and len(mobile_number) == 11 and mobile_number.startswith('09')):
             return render(request, 'booking/patient_login.html', {'error': 'شماره موبایل نامعتبر است. باید 11 رقم باشد و با 09 شروع شود.'})

        # --- OTP & SMS Sending Logic ---
        try:
            AMOOT_SMS_API_TOKEN = settings.AMOOT_SMS_API_TOKEN
            AMOOT_SMS_API_URL = settings.AMOOT_SMS_API_URL
            otp_code = str(random.randint(100000, 999999))
            request.session['otp_code_login'] = otp_code
            request.session['mobile_number_login'] = mobile_number
            request.session.set_expiry(300) # 5 minutes expiry for OTP

            payload = {
                'token': AMOOT_SMS_API_TOKEN,
                'Mobile': mobile_number,
                'PatternCodeID': 4018,
                'PatternValues': otp_code,
            }
            response = requests.post(AMOOT_SMS_API_URL, data=payload)
            return redirect('booking:verify_patient_login')

        except requests.exceptions.RequestException as e:
            print("خطا در ارسال پیامک:", e)
            return render(request, 'booking/patient_login.html', {'error': 'سیستم قادر به ارسال پیامک نمیباشد. لطفا با پشتیبانی تماس بگیرید.'})

    return render(request, 'booking/patient_login.html')

def verify_patient_login(request):
    """
    Verifies the OTP sent to the patient's mobile number and logs them in.
    """
    mobile_number = request.session.get('mobile_number_login')
    if not mobile_number:
        messages.error(request, 'اعتبار سنجی شما به پایان رسیده است. لطفا مجددا تلاش کنید.')
        return redirect('booking:patient_login')

    if request.method == 'POST':
        otp_from_user = request.POST.get('otp')
        otp_from_session = request.session.get('otp_code_login')

        if otp_from_user == otp_from_session:
            patient_user, created = User.objects.get_or_create(
                username=mobile_number,
                defaults={'user_type': 'PATIENT'}
            )
            login(request, patient_user)
            request.session.save()
            # Store the phone number for the dashboard to pick up
            request.session['verified_patient_phone'] = mobile_number

            for key in ['otp_code_login', 'mobile_number_login']:
                if key in request.session:
                    del request.session[key]

            messages.success(request, 'شما با موفقیت وارد شدید.')
            return redirect('booking:patient_dashboard')
        else:
            return render(request, 'booking/verify_patient_login.html', {'error': 'کد وارد شده صحیح نمی‌باشد.', 'mobile_number': mobile_number})

    return render(request, 'booking/verify_patient_login.html', {'mobile_number': mobile_number})

def patient_logout(request):
    """
    Logs the patient out.
    """
    logout(request)
    #messages.success(request, 'شما با موفقیت خارج شدید.')
    return redirect('booking:patient_login')


@login_required
def patient_dashboard(request):
    """
    Displays the patient's dashboard with their appointments.
    Allows patients to cancel their future appointments.
    """
    if request.user.user_type != 'PATIENT':
        return redirect('booking:doctor_list')
    appointments = []
    review_form = ReviewForm()  # Initialize the form

    if hasattr(request.user, 'user_type') and request.user.user_type == 'PATIENT':
        if request.method == 'POST':
            if 'appointment_id' in request.POST and 'rating' not in request.POST:
                appointment_id = request.POST.get('appointment_id')
                appointment_to_cancel = get_object_or_404(Appointment, pk=appointment_id, patient=request.user)
                if appointment_to_cancel.appointment_datetime.date() >= datetime.date.today():
                    appointment_to_cancel.status = 3
                    appointment_to_cancel.save()
                    messages.success(request, 'نوبت شما با موفقیت لغو شد.')
                else:
                    messages.error(request, 'شما نمی‌توانید نوبت‌های گذشته را لغو کنید.')
                return redirect('booking:patient_dashboard')

            elif 'appointment_id_review' in request.POST:
                review_form = ReviewForm(request.POST)  # Populate with POST data
                if review_form.is_valid():
                    appointment_id = request.POST.get('appointment_id_review')
                    appointment = get_object_or_404(Appointment, pk=appointment_id, patient=request.user)
                    review = review_form.save(commit=False)
                    review.appointment = appointment
                    review.save()
                    messages.success(request, 'نظر شما با موفقیت ثبت شد.')
                    return redirect('booking:patient_dashboard')
                # If form is invalid, execution continues and renders the page
                # with the populated, invalid form instance.

        appointments = Appointment.objects.filter(patient=request.user).select_related('doctor__user', 'doctor__specialty', 'review').order_by('-appointment_datetime')

    verified_phone = request.session.pop('verified_patient_phone', None)

    context = {
        'appointments': appointments,
        'page_title': 'نوبت‌های من',
        'today': datetime.date.today(),
        'verified_phone': verified_phone,
        'review_form': review_form, # This now contains the invalid form if submission failed
    }
    return render(request, 'booking/patient_dashboard.html', context)






from django.contrib.auth.views import LoginView
from django.shortcuts import redirect
from django.urls import reverse

class CustomLoginView(LoginView):
    template_name = 'booking/login.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.user_type == 'PATIENT':
            logout(request)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        user = self.request.user
        if user.is_authenticated:
            if user.user_type == 'DOCTOR':
                doctor_profile = getattr(user, 'doctor_profile', None)
                if doctor_profile and doctor_profile.financial_settings_completed:
                    return reverse('booking:daily_patients')
                else:
                    return reverse('booking:doctor_dashboard')
            elif user.user_type == 'SECRETARY':
                return reverse('booking:daily_patients')
            elif user.user_type == 'PATIENT':
                return reverse('booking:patient_dashboard')
        return reverse('booking:doctor_list')

@login_required
def export_expenses_to_excel(request):
    """
    Export expense balance report to Excel.
    """
    if not request.user.user_type == 'DOCTOR':
        return redirect('booking:doctor_list')

    doctor_profile = request.user.doctor_profile
    end_date = datetime.date.today()
    jalali_today = jdatetime.date.fromgregorian(date=end_date)
    start_of_year = jdatetime.date(jalali_today.year, 1, 1).togregorian()
    start_date = start_of_year

    expenses = DailyExpense.objects.filter(
        doctor=doctor_profile,
        date__range=[start_date, end_date],
        amount__gt=0
    ).values('description').annotate(
        count=Count('id'),
        total_amount=Sum('amount'),
        average_amount=Avg('amount')
    ).order_by('-total_amount')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expenses'

    columns = [
        'شرح',
        'تعداد',
        'مجموع',
        'میانگین',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for expense in expenses:
        row_num += 1
        row = [
            expense['description'],
            expense['count'],
            expense['total_amount'],
            expense['average_amount'],
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

